import os

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from tkinter import filedialog
import tkinter

env = Environment(loader=FileSystemLoader('templates'))


def select_mode(mode):
    start_row = 3
    if mode == 1:
        # contingent
        keys = ['code', 'name', 'level', 'form', 'BF', 'BFF', 'BR', 'BRF', 'BM', 'BMF', 'P', 'PF', 'All']
        table = "contingent_table.html"
    elif mode == 2:
        # vacant
        keys = ['code', 'name', 'level', 'prof', 'course', 'form', 'BFVacant', 'BRVacant',
                'BMVacant', 'PVacant']
        table = "vacant_table.html"
    elif mode == 3:
        # priem
        keys = ['code', 'name', 'level', 'form', 'BF', 'BR', 'BM', 'P', 'score']
        table = "priem_table.html"
    elif mode == 4:
        # perevod
        keys = ['code', 'name', 'level', 'form', 'NO', 'NT', 'NR', 'NE']
        table = "perevod_table.html"
        start_row = 2
    elif mode == 5:
        # praktika
        keys = ['code', 'name', 'year', 'profile', 'form', 'subject', 'technology',
                'PRU', 'PRP', 'PRD']
        table = "praktika_table.html"
    elif mode == 6:
        # akkred
        keys = ['code', 'name', 'prof', 'level', 'form', 'learningTerm', 'dateEnd',
                'language', 'discipline', 'practice', 'eos']
        table = "akkred_table.html"
        start_row = 2
    elif mode == 7:
        # obrazovanie
        keys = ['code', 'name', 'level', 'prof', 'form', 'OOP', 'plan', 'ann', 'rpd', 'shl',
                'met', 'prac', 'eos']
        table = "obrazovanie_table.html"
        start_row = 2
    elif mode == 8:
        # NIR
        keys = ['code', 'name', 'perechen', 'prof', 'level', 'naprav', 'result', 'base']
        table = "nir_table.html"
        start_row = 2
    elif mode == 9:
        # prepodavateli
        keys = ['fio', 'post', 'level', 'tqual', 'equal', 'degree', 'academic_stat',
                'general_exp', 'special_exp', 'prof_dev', 'discipline', 'teach_areas', 'honors']
        # table = "teachers_detail.html"
        table = "teachers_detail_all_in_one.html"
        start_row = 2
    elif mode == 10:
        # mezd
        keys = ['id', 'country', 'name', 'dog']
        table = "mezd_table.html"
        start_row = 2
    elif mode == 11:
        # obr standart
        keys = ['code', 'name', 'level', 'doc3', 'doc3plus', 'fedTreb', 'locStand', 'locTreb']
        table = "standart_table.html"
        start_row = 2
    elif mode == 12:
        # trudoustr
        keys = ['code', 'name', 'prof', 'form', 'v1', 't1']
        table = "vypusk_table.html"
    elif mode == 13:
        # cab prac
        keys = ['address', 'name', 'osn']
        table = "prCab_table.html"
        start_row = 2
    elif mode == 14:
        # cab uch
        keys = ['address', 'name', 'osn']
        table = "uCab_table.html"
        start_row = 2
    else:
        raise Exception
    return table, keys, start_row


def html_tbl(template_name, keys, sheet, start_from):
    set_ = []
    dictionary = {}
    current_template = env.get_template(template_name)
    for i in range(start_from, sheet.max_row + 1):
        for j in range(len(keys)):
            if not(sheet[i][j].value is None):
                if sheet[i][j].hyperlink:
                    link = sheet[i][j].hyperlink.display
                    dictionary[keys[j]] = f'<a href="{link}">{sheet[i][j].value}</a>'
                else:
                    dictionary[keys[j]] = sheet[i][j].value
            else:
                dictionary[keys[j]] = '–'
        set_.append(dictionary.copy())
    return current_template.render(sheet=set_)


def get_file_name(file_path, file):
    if os.path.isfile(os.path.join(file_path, file)):
        i = 1
        temp_name, extension = str.rsplit(file, '.', 1)
        while True:
            file = temp_name + '(' + str(i) + ').' + extension
            i += 1
            if not os.path.isfile(os.path.join(file_path, file)):
                break
        return file
    else:
        return file


if __name__ == '__main__':
    try:
        tkinter.Tk().withdraw()
        path_to_file = tkinter.filedialog.askopenfilename()
        print("Выберите режим (только цифра): \n 1. Численность \n 2. Вакантные места "
              "\n 3. Результаты приема \n 4. Перевод \n 5. Практика \n 6. Аккредитация "
              "\n 7. Образование \n 8. Результаты научной деятельности \n 9. Преподы "
              "\n 10. Международное сотрудничество \n 11. Обр. стандарты "
              "\n 12. Трудоустройство \n 13. Кабинеты практика \n 14. Кабинеты учебные")

        current_mode = 15
        while current_mode > 14 or current_mode == 0:
            current_mode = int(input(">> "))

        book = load_workbook(filename=path_to_file)
        worksheets = book.worksheets
        sheet = worksheets[0]

        temp_path = path_to_file.rsplit('/', 1)
        dir_path = temp_path[0] + '/done/'
        file_name = temp_path[1].replace('xlsx', 'html')
        file_name = get_file_name(dir_path, file_name)
        final_path = os.path.join(dir_path, file_name)

        if not os.path.exists(dir_path):
            os.mkdir(dir_path)

        final_tbl = html_tbl(select_mode(current_mode)[0], select_mode(current_mode)[1], sheet, select_mode(current_mode)[2])

        with open(final_path, "w", encoding="utf-8") as final_file:
            final_file.write(final_tbl)
            print('\033[32m' + 'OK. Saved to: ', final_path)

    except Exception as e:
        print(e)
